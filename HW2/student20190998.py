#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

total = []
row_id = 1 
people = 0
for row in ws:
	sum_v = 0
	if row_id != 1:
		sum_v += ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		
		ws.cell(row = row_id, column = 7).value = sum_v
		
		total.append(ws.cell(row = row_id, column = 7).value)
		
		people += 1
	row_id += 1
	

total.sort(reverse=True)

Aplus = []
Azero = []
Bplus = []
Bzero = []
Cplus = []
Czero = []
Fzero = []
n = 1
for t in total:
	if t < 40:
		Fzero.append(t)
	elif n <= people * 15 // 100:
		Aplus.append(t)
	elif n <= people * 30 // 100:
		Azero.append(t)
	elif n <= people * 50 // 100:
		Bplus.append(t)
	elif n <= people * 70 // 100:
		Bzero.append(t)
	else:
		if n <= people * 85 // 100:
			Cplus.append(t)
		else:
			Czero.append(t)
	n += 1

row_id = 1
for row in ws:
	if row_id != 1:
		t = ws.cell(row = row_id, column = 7).value
		if t in Aplus:
			ws.cell(row = row_id, column = 8).value = "A+"
		elif t in Azero:
			ws.cell(row = row_id, column = 8).value = "A0"
		elif t in Bplus:
			ws.cell(row = row_id, column = 8).value = "B+"
		elif t in Bzero:
			ws.cell(row = row_id, column = 8).value = "B0"
		elif t in Cplus:
			ws.cell(row = row_id, column = 8).value = "C+"
		elif t in Czero:
			ws.cell(row = row_id, column = 8).value = "C0"
		else:
			ws.cell(row = row_id, column = 8).value = "F"
	row_id += 1
wb.save("student.xlsx")
